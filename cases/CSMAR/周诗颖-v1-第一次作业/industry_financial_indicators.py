import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.marker import Marker

# 创建保存结果的目录
os.makedirs("分析结果", exist_ok=True)

# 尝试设置中文字体
try:
    plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei']
    plt.rcParams['axes.unicode_minus'] = False
except:
    print("警告: 未能正确设置中文字体，图表中的中文可能无法正确显示")

# 读取Excel文件
print("正在读取数据文件...")
try:
    data = pd.read_excel("金融数据分析.xlsx")
    print(f"数据读取成功! 共有{data.shape[0]}行, {data.shape[1]}列")
except Exception as e:
    print(f"读取数据时出错: {e}")
    exit(1)

# 检查必要的财务指标列是否存在
required_indicators = ['SLoan', 'LLoan', 'Lev', 'Cash', 'ROA', 'ROE']
available_indicators = [col for col in required_indicators if col in data.columns]
missing_indicators = [col for col in required_indicators if col not in data.columns]

print(f"可用的财务指标: {available_indicators}")
if missing_indicators:
    print(f"警告: 数据中缺少以下财务指标: {missing_indicators}")
    print("分析将继续，但结果可能不完整")

# 处理年份列
# 第一行可能是中文标题，跳过
if isinstance(data['Year'].iloc[0], str) and data['Year'].iloc[0] == '年份':
    data = data.iloc[1:].reset_index(drop=True)

# 确保Year列是数值类型
try:
    data['Year'] = pd.to_numeric(data['Year'], errors='coerce')
    data = data.dropna(subset=['Year'])  # 删除年份为空的行
    data['year'] = data['Year'].astype(int)  # 创建一个年份的整数列
except Exception as e:
    print(f"转换年份时出错: {e}")
    print("尝试手动检查年份数据...")
    print(data['Year'].head(10))  # 打印前10条年份数据进行检查

# 行业代码映射
industry_mapping = {
    'C': '制造业',
    'D': '电力、热力、燃气及水生产和供应业',
    'G': '交通运输业',
    'E': '建筑业',
    'K': '房地产业',
    'F': '批发和零售业',
    'J': '金融业'
}

# 处理行业代码列
def extract_industry_code(code):
    if pd.isna(code):
        return np.nan
    match = re.match(r'([A-Z])', str(code))
    if match:
        return match.group(1)
    return np.nan

data['industry_code'] = data['Industry'].apply(extract_industry_code)

# 选择要分析的行业
selected_industries = ['C', 'D', 'G', 'E', 'K']  # 制造业、电力、交通、建筑、房地产
print(f"选择分析的行业: {[industry_mapping.get(code, '未知') for code in selected_industries]}")

# 筛选所选行业的数据
data_filtered = data[data['industry_code'].isin(selected_industries)]

# 选择要分析的年份（1999年之后的奇数年份，以及2024年）
selected_years = list(range(1999, 2024, 2)) + [2024]
print(f"选择分析的年份: {selected_years}")

# 筛选所选年份的数据
data_filtered = data_filtered[data_filtered['year'].isin(selected_years)]

# 检查数据量是否足够
year_counts = data_filtered['year'].value_counts().sort_index()
print("\n各年份数据量:")
for year, count in year_counts.items():
    print(f"  {year}年: {count}条记录")

if data_filtered.empty:
    print("错误: 筛选后没有数据可供分析")
    exit(1)

# 检查是否有所有所需的指标
for indicator in required_indicators:
    if indicator in data_filtered.columns:
        missing_values = data_filtered[indicator].isna().sum()
        total_values = len(data_filtered)
        if missing_values > 0:
            missing_percent = missing_values / total_values * 100
            print(f"警告: 指标'{indicator}'有{missing_values}条缺失值 ({missing_percent:.2f}%)")

# 为每个指标创建数据透视表，计算每个行业每年的平均值
industry_year_tables = {}
for indicator in available_indicators:
    # 计算每个行业每年的平均值
    pivot = data_filtered.pivot_table(
        values=indicator,
        index='industry_code',
        columns='year',
        aggfunc='mean'
    )
    
    # 将行业代码替换为行业名称
    pivot.index = pivot.index.map(lambda x: f"{industry_mapping.get(x, '未知')} ({x})")
    
    # 保存数据透视表
    industry_year_tables[indicator] = pivot
    
    # 保存为CSV文件
    pivot.to_csv(f"分析结果/{indicator}_行业年份分析.csv", encoding='utf-8-sig')
    print(f"已保存{indicator}指标分析结果到CSV文件")

# 创建一个Excel文件保存所有表格
excel_file = "分析结果/行业财务指标分析.xlsx"
wb = Workbook()
ws_overview = wb.active
ws_overview.title = "总览"

# 设置单元格样式函数
def set_header_style(cell):
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

def set_data_style(cell, is_header=False):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    if is_header:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

# 在总览表中创建行业财务指标概览
overview_title = "行业财务指标平均值分析 (1999-2024)"
ws_overview['A1'] = overview_title
ws_overview.merge_cells('A1:I1')
ws_overview['A1'].font = Font(bold=True, size=14)
ws_overview['A1'].alignment = Alignment(horizontal="center", vertical="center")

# 添加列标题
headers = ['行业', '短期负债率\n(SLoan)', '长期负债率\n(LLoan)', '总负债率\n(Lev)', '现金比率\n(Cash)', '总资产回报率\n(ROA)', '股本回报率\n(ROE)']
for col, header in enumerate(headers, start=1):
    cell = ws_overview.cell(row=3, column=col)
    cell.value = header
    set_header_style(cell)
    # 设置列宽
    ws_overview.column_dimensions[get_column_letter(col)].width = 15

# 添加行业数据
row = 4
for industry_code in selected_industries:
    industry_name = industry_mapping.get(industry_code, '未知')
    ws_overview.cell(row=row, column=1).value = f"{industry_name} ({industry_code})"
    set_data_style(ws_overview.cell(row=row, column=1), is_header=True)
    
    col = 2
    for indicator in ['SLoan', 'LLoan', 'Lev', 'Cash', 'ROA', 'ROE']:
        if indicator in available_indicators:
            # 获取该行业该指标的所有年份平均值
            pivot = industry_year_tables[indicator]
            industry_label = f"{industry_name} ({industry_code})"
            if industry_label in pivot.index:
                avg_value = pivot.loc[industry_label].mean()
                ws_overview.cell(row=row, column=col).value = avg_value
                ws_overview.cell(row=row, column=col).number_format = '0.0000'
                set_data_style(ws_overview.cell(row=row, column=col))
        col += 1
    row += 1

# 为每个指标创建独立的工作表
indicator_descriptions = {
    'SLoan': '短期负债率 - 反映企业短期偿债压力',
    'LLoan': '长期负债率 - 反映企业长期融资结构',
    'Lev': '总负债率 - 反映企业整体杠杆水平',
    'Cash': '现金比率 - 反映企业流动性状况',
    'ROA': '总资产回报率 - 反映企业资产使用效率',
    'ROE': '股本回报率 - 反映企业股东回报水平'
}

for indicator in available_indicators:
    # 创建新表
    ws = wb.create_sheet(title=indicator)
    
    # 添加标题
    title = f"{indicator_descriptions.get(indicator, indicator)} 行业分析 (1999-2024)"
    ws['A1'] = title
    ws.merge_cells(f'A1:{get_column_letter(len(selected_years)+2)}1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加说明
    ws['A2'] = f"数据显示了各行业{indicator}指标在不同年份的平均值"
    ws.merge_cells(f'A2:{get_column_letter(len(selected_years)+2)}2')
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加列标题（年份）
    ws['A3'] = "行业"
    set_header_style(ws['A3'])
    for i, year in enumerate(sorted(selected_years)):
        cell = ws.cell(row=3, column=i+2)
        cell.value = year
        set_header_style(cell)
    
    # 添加"变化率"列
    cell = ws.cell(row=3, column=len(selected_years)+2)
    cell.value = "变化率 (%)"
    set_header_style(cell)
    
    # 添加行业数据
    pivot = industry_year_tables[indicator]
    for i, industry_code in enumerate(selected_industries):
        industry_name = industry_mapping.get(industry_code, '未知')
        industry_label = f"{industry_name} ({industry_code})"
        
        # 添加行业名称
        row = i + 4
        cell = ws.cell(row=row, column=1)
        cell.value = industry_label
        set_data_style(cell, is_header=True)
        
        # 添加各年份数据
        if industry_label in pivot.index:
            for j, year in enumerate(sorted(selected_years)):
                col = j + 2
                if year in pivot.columns:
                    value = pivot.loc[industry_label, year]
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.number_format = '0.0000'
                    set_data_style(cell)
            
            # 计算变化率（最后一年与第一年相比）
            first_year = sorted(selected_years)[0]
            last_year = sorted(selected_years)[-1]
            if first_year in pivot.columns and last_year in pivot.columns:
                first_value = pivot.loc[industry_label, first_year]
                last_value = pivot.loc[industry_label, last_year]
                if first_value != 0 and not pd.isna(first_value) and not pd.isna(last_value):
                    change_pct = (last_value - first_value) / first_value * 100
                    cell = ws.cell(row=row, column=len(selected_years)+2)
                    cell.value = change_pct
                    cell.number_format = '0.00'
                    set_data_style(cell)
                    
                    # 根据变化率设置颜色
                    if change_pct > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif change_pct < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 调整列宽
    ws.column_dimensions['A'].width = 30
    for i in range(len(selected_years) + 2):
        col_letter = get_column_letter(i + 2)
        ws.column_dimensions[col_letter].width = 15
    
    # 添加柱状图，显示最新年份的各行业数据比较
    latest_year = sorted(selected_years)[-1]
    if latest_year in pivot.columns:
        chart_sheet = wb.create_sheet(title=f"{indicator}_图表")
        
        # 创建柱状图
        bar_chart = BarChart()
        bar_chart.title = f"{latest_year}年各行业{indicator}指标比较"
        bar_chart.style = 2
        bar_chart.x_axis.title = "行业"
        bar_chart.y_axis.title = indicator
        
        data = Reference(ws, min_col=len(selected_years)+1, min_row=3, max_row=3+len(selected_industries), max_col=len(selected_years)+1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=3+len(selected_industries))
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        
        chart_sheet.add_chart(bar_chart, "A1")
        
        # 创建时序图，显示各行业指标随时间的变化
        line_chart = LineChart()
        line_chart.title = f"各行业{indicator}指标随时间变化趋势"
        line_chart.style = 2
        line_chart.x_axis.title = "年份"
        line_chart.y_axis.title = indicator
        
        # 设置类别（年份）
        cats = Reference(ws, min_col=2, max_col=len(selected_years)+1, min_row=3, max_row=3)
        
        # 添加每个行业的数据系列
        for i in range(len(selected_industries)):
            data = Reference(ws, min_col=2, max_col=len(selected_years)+1, min_row=4+i, max_row=4+i)
            series = line_chart.series[i] if i < len(line_chart.series) else None
            if series is None:
                line_chart.add_data(data, titles_from_data=False)
                line_chart.series[-1].marker = Marker('circle')
                # 设置系列名称
                industry_label = ws.cell(row=4+i, column=1).value
                line_chart.series[-1].title = industry_label
            else:
                series.values = data
        
        line_chart.set_categories(cats)
        
        chart_sheet.add_chart(line_chart, "A20")

# 保存Excel文件
try:
    wb.save(excel_file)
    print(f"\n已保存分析结果到: {excel_file}")
except Exception as e:
    print(f"保存Excel文件时出错: {e}")

# 执行财务指标分析总结
print("\n== 财务指标分析总结 ==")

# 债务结构分析
print("\n1. 债务结构分析 (SLoan, LLoan, Lev):")
for industry_code in selected_industries:
    industry_name = industry_mapping.get(industry_code, '未知')
    industry_label = f"{industry_name} ({industry_code})"
    
    sloan_pivot = industry_year_tables.get('SLoan', pd.DataFrame())
    lloan_pivot = industry_year_tables.get('LLoan', pd.DataFrame())
    lev_pivot = industry_year_tables.get('Lev', pd.DataFrame())
    
    if industry_label in lev_pivot.index:
        avg_lev = lev_pivot.loc[industry_label].mean()
        recent_lev = lev_pivot.loc[industry_label][lev_pivot.columns[-1]]
        
        debt_structure = []
        if industry_label in sloan_pivot.index and industry_label in lloan_pivot.index:
            avg_sloan = sloan_pivot.loc[industry_label].mean()
            avg_lloan = lloan_pivot.loc[industry_label].mean()
            
            if avg_lev > 0:
                sloan_ratio = avg_sloan / avg_lev * 100
                lloan_ratio = avg_lloan / avg_lev * 100
                debt_structure.append(f"短期负债占比{sloan_ratio:.2f}%，长期负债占比{lloan_ratio:.2f}%")
        
        # 负债率评估
        if avg_lev > 0.7:
            leverage_assessment = "高负债水平"
        elif avg_lev > 0.5:
            leverage_assessment = "中等负债水平"
        else:
            leverage_assessment = "较低负债水平"
        
        print(f"  {industry_label}: 平均负债率{avg_lev:.4f}, 最新负债率{recent_lev:.4f}, {leverage_assessment}" + (f", {debt_structure[0]}" if debt_structure else ""))

# 盈利能力分析
print("\n2. 盈利能力分析 (ROA, ROE):")
for industry_code in selected_industries:
    industry_name = industry_mapping.get(industry_code, '未知')
    industry_label = f"{industry_name} ({industry_code})"
    
    roa_pivot = industry_year_tables.get('ROA', pd.DataFrame())
    roe_pivot = industry_year_tables.get('ROE', pd.DataFrame())
    
    if industry_label in roa_pivot.index and industry_label in roe_pivot.index:
        avg_roa = roa_pivot.loc[industry_label].mean()
        avg_roe = roe_pivot.loc[industry_label].mean()
        
        # ROA评估
        if avg_roa > 0.10:
            roa_assessment = "优异的资产利用效率"
        elif avg_roa > 0.05:
            roa_assessment = "良好的资产利用效率"
        elif avg_roa > 0:
            roa_assessment = "一般的资产利用效率"
        else:
            roa_assessment = "较低的资产利用效率"
        
        # ROE评估
        if avg_roe > 0.15:
            roe_assessment = "优异的股东回报"
        elif avg_roe > 0.10:
            roe_assessment = "良好的股东回报"
        elif avg_roe > 0:
            roe_assessment = "一般的股东回报"
        else:
            roe_assessment = "较低的股东回报"
        
        print(f"  {industry_label}: 平均ROA {avg_roa:.4f} ({roa_assessment}), 平均ROE {avg_roe:.4f} ({roe_assessment})")
        
        # 财务杠杆效应
        if avg_roa > 0 and avg_roe > 0 and avg_roe > avg_roa:
            leverage_effect = (avg_roe - avg_roa) / avg_roa * 100
            print(f"    财务杠杆效应: +{leverage_effect:.2f}% (ROE > ROA, 正向财务杠杆)")
        elif avg_roa > 0 and avg_roe > 0 and avg_roe < avg_roa:
            leverage_effect = (avg_roe - avg_roa) / avg_roa * 100
            print(f"    财务杠杆效应: {leverage_effect:.2f}% (ROE < ROA, 负向财务杠杆)")

# 流动性分析
print("\n3. 流动性分析 (Cash):")
for industry_code in selected_industries:
    industry_name = industry_mapping.get(industry_code, '未知')
    industry_label = f"{industry_name} ({industry_code})"
    
    cash_pivot = industry_year_tables.get('Cash', pd.DataFrame())
    
    if industry_label in cash_pivot.index:
        avg_cash = cash_pivot.loc[industry_label].mean()
        
        # 现金比率评估
        if avg_cash > 0.3:
            cash_assessment = "充裕的流动性"
        elif avg_cash > 0.2:
            cash_assessment = "良好的流动性"
        elif avg_cash > 0.1:
            cash_assessment = "一般的流动性"
        else:
            cash_assessment = "较低的流动性"
        
        print(f"  {industry_label}: 平均现金比率 {avg_cash:.4f} ({cash_assessment})")

# 指标变化趋势分析
print("\n4. 各指标变化趋势分析 (1999-2024):")
for indicator in available_indicators:
    print(f"\n  {indicator} ({indicator_descriptions.get(indicator, '')}) 变化趋势:")
    pivot = industry_year_tables[indicator]
    
    for industry_code in selected_industries:
        industry_name = industry_mapping.get(industry_code, '未知')
        industry_label = f"{industry_name} ({industry_code})"
        
        if industry_label in pivot.index:
            # 计算25年间的变化率
            first_year = sorted(selected_years)[0]
            last_year = sorted(selected_years)[-1]
            
            if first_year in pivot.columns and last_year in pivot.columns:
                first_value = pivot.loc[industry_label, first_year]
                last_value = pivot.loc[industry_label, last_year]
                
                if not pd.isna(first_value) and not pd.isna(last_value) and first_value != 0:
                    change_pct = (last_value - first_value) / first_value * 100
                    direction = "↑" if change_pct > 0 else "↓"
                    
                    # 变化趋势描述
                    if abs(change_pct) < 10:
                        trend_desc = "基本保持稳定"
                    elif abs(change_pct) < 50:
                        trend_desc = "有所" + ("增长" if change_pct > 0 else "下降")
                    elif abs(change_pct) < 100:
                        trend_desc = "显著" + ("增长" if change_pct > 0 else "下降")
                    else:
                        trend_desc = "大幅" + ("增长" if change_pct > 0 else "下降")
                    
                    print(f"    {industry_label}: {change_pct:.2f}% {direction} ({trend_desc})")

print("\n分析完成!") 