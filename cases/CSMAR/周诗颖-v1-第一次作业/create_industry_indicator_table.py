import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建保存结果的目录
os.makedirs("分析结果", exist_ok=True)

# 尝试设置中文字体
try:
    plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei']
    plt.rcParams['axes.unicode_minus'] = False
except:
    print("警告: 未能正确设置中文字体，图表中的中文可能无法正确显示")

# 读取Excel文件
print("正在读取数据文件...")
try:
    data = pd.read_excel("金融数据分析.xlsx")
    print("数据读取成功！")
except Exception as e:
    print(f"读取数据时出错: {e}")
    exit(1)

# 显示数据的基本信息
print(f"数据行数: {data.shape[0]}, 数据列数: {data.shape[1]}")

# 检查必要的列是否存在
required_indicators = ['SLoan', 'LLoan', 'Lev', 'Cash', 'ROA', 'ROE']
required_columns = ['Year', 'Industry'] + required_indicators
missing_columns = [col for col in required_columns if col not in data.columns]
if missing_columns:
    print(f"错误: 数据中缺少以下必要的列: {missing_columns}")
    exit(1)

# 处理年份列
# 第一行可能是中文标题，跳过
if data['Year'].iloc[0] == '年份':
    data = data.iloc[1:].reset_index(drop=True)

# 确保Year列是数值类型
try:
    data['Year'] = pd.to_numeric(data['Year'], errors='coerce')
    data = data.dropna(subset=['Year'])  # 删除年份为空的行
    data['year'] = data['Year'].astype(int)  # 创建一个年份的整数列
except Exception as e:
    print(f"转换年份时出错: {e}")
    print("尝试手动检查年份数据...")
    print(data['Year'].head(10))  # 打印前10条年份数据进行检查

# 行业代码映射
industry_mapping = {
    'C': '制造业',
    'D': '电力、热力、燃气及水生产和供应业',
    'G': '交通运输业',
    'E': '建筑业',
    'K': '房地产业',
    'F': '批发和零售业',
    'J': '金融业'
}

# 处理行业代码列
# 提取证监会行业分类主代码 (假设格式为字母+数字，如C39)
def extract_industry_code(code):
    if pd.isna(code):
        return np.nan
    match = re.match(r'([A-Z])', str(code))
    if match:
        return match.group(1)
    return np.nan

data['industry_code'] = data['Industry'].apply(extract_industry_code)
print("\n行业代码分布:")
print(data['industry_code'].value_counts())

# 筛选所需的行业
target_industries = list(industry_mapping.keys())
data_filtered = data[data['industry_code'].isin(target_industries)]

# 选定奇数年份
selected_years = list(range(1999, 2024, 2))  # 从1999年开始，每隔2年选择一个年份
print(f"\n选定的年份: {selected_years}")

# 筛选所选年份的数据
data_selected_years = data_filtered[data_filtered['year'].isin(selected_years)]

# 检查是否有足够的数据
year_counts = data_selected_years['year'].value_counts().sort_index()
print("\n各年份数据量:")
print(year_counts)

# 确认所有指标都存在
print("\n检查所需指标是否存在:")
for indicator in required_indicators:
    if indicator in data_selected_years.columns:
        print(f"  {indicator}: 存在")
    else:
        print(f"  {indicator}: 不存在")
        required_indicators.remove(indicator)

# 为每个指标创建行业-年份交叉表
indicator_tables = {}
for indicator in required_indicators:
    # 计算每个行业每年的平均值
    avg_by_industry_year = data_selected_years.groupby(['industry_code', 'year'])[indicator].mean().reset_index()
    
    # 转换为行业为行，年份为列的表格
    pivot_table = avg_by_industry_year.pivot(index='industry_code', columns='year', values=indicator)
    
    # 用行业名称替换行业代码
    pivot_table.index = pivot_table.index.map(lambda x: f"{industry_mapping.get(x, '未知')} ({x})")
    
    # 保存到字典
    indicator_tables[indicator] = pivot_table
    
    # 保存到CSV
    pivot_table.to_csv(f"分析结果/{indicator}_行业年度平均值.csv", encoding='utf-8-sig')
    print(f"已保存 {indicator} 行业年度平均值至 CSV 文件")

# 创建Excel文件保存所有表格
excel_file = "分析结果/行业财务指标分析.xlsx"
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    # 先创建一个总览表
    overview = pd.DataFrame(index=indicator_tables[required_indicators[0]].index)
    
    # 对每个指标，计算每个行业的平均值
    for indicator in required_indicators:
        indicator_avg = indicator_tables[indicator].mean(axis=1)
        overview[f"{indicator}_平均值"] = indicator_avg
    
    # 保存总览表
    overview.to_excel(writer, sheet_name="指标总览")
    
    # 保存每个指标的详细表格
    for indicator in required_indicators:
        indicator_tables[indicator].to_excel(writer, sheet_name=indicator)

print(f"\n已将所有指标表格保存至 {excel_file}")

# 使用openpyxl美化Excel表格
wb = Workbook()
wb = writer.book

# 定义样式
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# 对每个工作表应用样式
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # 调整列宽
    for column in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 15
    
    # 应用标题样式
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # 应用数据单元格样式
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col > 1:  # 数值列
                cell.alignment = Alignment(horizontal="right")

# 保存样式
wb.save(excel_file)

# 进行指标分析
print("\n== 行业财务指标分析 ==")

# 1. 债务结构分析 (SLoan, LLoan, Lev)
print("\n债务结构分析:")
debt_indicators = ['SLoan', 'LLoan', 'Lev']
debt_indicators = [ind for ind in debt_indicators if ind in required_indicators]

if debt_indicators:
    # 计算平均值并排序
    debt_avg = {}
    for indicator in debt_indicators:
        debt_avg[indicator] = indicator_tables[indicator].mean(axis=1).sort_values(ascending=False)
    
    # 打印排名
    for indicator in debt_indicators:
        print(f"\n{indicator} 平均值排名:")
        for industry, value in debt_avg[indicator].items():
            print(f"  {industry}: {value:.4f}")

# 2. 盈利能力分析 (ROA, ROE)
print("\n盈利能力分析:")
profit_indicators = ['ROA', 'ROE']
profit_indicators = [ind for ind in profit_indicators if ind in required_indicators]

if profit_indicators:
    # 计算平均值并排序
    profit_avg = {}
    for indicator in profit_indicators:
        profit_avg[indicator] = indicator_tables[indicator].mean(axis=1).sort_values(ascending=False)
    
    # 打印排名
    for indicator in profit_indicators:
        print(f"\n{indicator} 平均值排名:")
        for industry, value in profit_avg[indicator].items():
            print(f"  {industry}: {value:.4f}")

# 3. 流动性分析 (Cash)
print("\n流动性分析:")
if 'Cash' in required_indicators:
    cash_avg = indicator_tables['Cash'].mean(axis=1).sort_values(ascending=False)
    
    print("\nCash 平均值排名:")
    for industry, value in cash_avg.items():
        print(f"  {industry}: {value:.4f}")

# 4. 趋势分析 - 计算每个指标的变化率
print("\n各指标变化趋势分析:")
for indicator in required_indicators:
    print(f"\n{indicator} 变化率分析:")
    table = indicator_tables[indicator]
    
    # 获取第一年和最后一年
    years = sorted(table.columns)
    if len(years) >= 2:
        first_year = years[0]
        last_year = years[-1]
        
        # 计算变化率
        pct_changes = ((table[last_year] - table[first_year]) / table[first_year] * 100).sort_values(ascending=False)
        
        print(f"从{first_year}年到{last_year}年的变化率:")
        for industry, change in pct_changes.items():
            if not np.isnan(change):
                direction = "↑" if change > 0 else "↓"
                print(f"  {industry}: {change:.2f}% {direction}")

print("\n分析完成!") 